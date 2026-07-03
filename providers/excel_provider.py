import os

from .base_provider import BaseProvider

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ExcelProvider(BaseProvider):

    name = "excel_provider"

    def _request(self, xlsx_path: str, **kwargs):
        return self._parse_xlsx(xlsx_path)

    def _parse_xlsx(self, xlsx_path: str):
        if not os.path.isfile(xlsx_path):
            return {
                "text": "",
                "tables": [],
                "meta": {
                    "source": xlsx_path,
                    "sheets": 0,
                    "mock": False,
                    "error": f"Excel file not found: {xlsx_path}",
                },
            }

        if openpyxl is None:
            return {
                "text": "Unable to parse Excel. Please install openpyxl.",
                "tables": [],
                "meta": {
                    "source": xlsx_path,
                    "sheets": 0,
                    "mock": False,
                    "error": "openpyxl is not installed",
                },
            }

        try:
            workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as exc:
            return {
                "text": "Unable to parse Excel file.",
                "tables": [],
                "meta": {
                    "source": xlsx_path,
                    "sheets": 0,
                    "mock": False,
                    "error": f"Excel parse error: {exc}",
                },
            }

        tables = []
        raw_text = ""

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = []
            sheet_lines = [f"--- Sheet: {sheet_name} ---"]
            for row in sheet.iter_rows(values_only=True):
                row_values = ["" if cell is None else str(cell) for cell in row]
                if any(cell != "" for cell in row_values):
                    rows.append(row_values)
                    sheet_lines.append("\t".join(row_values))
            if rows:
                tables.append({
                    "sheet": sheet_name,
                    "rows": rows,
                })
            raw_text += "\n" + "\n".join(sheet_lines)

        return {
            "text": raw_text.strip(),
            "tables": tables,
            "meta": {
                "source": xlsx_path,
                "sheets": len(workbook.sheetnames),
                "mock": False,
                "method": "openpyxl",
            },
        }
